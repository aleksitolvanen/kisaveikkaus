#!/usr/bin/env python3
"""Tuo veikkaukset tai oikeat tulokset kisaveikkaus-Excelistä JSONiksi.

Käyttö:
    python tools/import-excel.py <xlsx> <tournamentId> --mode predictions
    python tools/import-excel.py <xlsx> <tournamentId> --mode results
    python tools/import-excel.py <xlsx> <tournamentId> --mode all

Kirjoittaa kohteeseen data/<tournamentId>/{tournament,predictions,results}.json.

Sama "Tulokset"-välilehden rakenne toistuu vuodesta toiseen (ryhmä jakaa saman
Excel-pohjan), joten skripti toimii myös tulevien turnausten Exceleille. Rakenne
tunnistetaan pääosin automaattisesti (lohko-otsikot, ottelut XXX-YYY-regexillä,
Sikajengi-/Maalintekijä-rivit labelista). Cup-vaiheen rivivälit ovat
eksplisiittisessä LAYOUT-kartassa. Jos tuleva pohja eroaa, päivitä LAYOUT.

Vaatii: openpyxl  (pip install openpyxl)
"""
import argparse
import json
import os
import re
import sys
from datetime import datetime, timedelta, timezone

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    sys.exit("Tarvitaan openpyxl:  pip install openpyxl")

SHEET = "Tulokset"
RESULT_COL = column_index_from_string("C")  # "TULOS" – oikeat tulokset
FIRST_PRED_COL = column_index_from_string("E")  # ensimmäinen veikkaajan sarake
MATCH_RE = re.compile(r"^([A-Za-z]{2,4})-([A-Za-z]{2,4})$")

# Excel-pohjan joukkuekoodit -> FIFA:n viralliset koodit. JSON-data käyttää aina
# FIFA-koodeja; vain Excel poikkeaa. Kartta normalisoi kaikki koodikentät
# (ottelut, cup-veikkaukset, sikajengi) tuonnissa.
CODE = {"CUR": "CUW", "ICV": "CIV", "DRC": "COD", "SPA": "ESP", "SWI": "SUI",
        "SER": "SRB", "ROM": "ROU", "KOL": "COL"}

# Maalintekijänimet yhtenäisiksi: pisteytys vertaa results.goals-avaimia
# täsmälleen, joten sama pelaaja tarvitsee saman kirjoitusasun kaikilla.
GOALSCORER_FIX = {"MBappe": "Mbappe", "Kai Havertzin": "Havertz"}


def norm(code):
    c = str(code).strip().upper()
    return CODE.get(c, c)

# Template "mm-2026": cup-vaiheen rivivälit (inclusive).
LAYOUT = {
    # Veikkaukset ja oikeat tulokset luetaan samoilta riveiltä: veikkaajan
    # sarakkeesta (E,G,...) ja oikeiden tulosten C-sarakkeesta. Ylläpitäjä
    # täyttää oikeat jatkoonpääsijät C-sarakkeeseen samoille riveille.
    "cup":       {"r16": (93, 108), "qf": (110, 117), "sf": (119, 122),
                  "final": (124, 125), "champion": (127, 127)},
    "cupMeta":   {"r16": ("16 joukkoon", 2), "qf": ("Puolivälierä", 4),
                  "sf": ("Välierä", 8), "final": ("Finaali", 15),
                  "champion": ("Mestari", 30)},
}
ROUND_ORDER = ["r16", "qf", "sf", "final", "champion"]


def cellval(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def find_label_row(ws, needle, maxrow=200):
    needle = needle.lower()
    for r in range(1, maxrow + 1):
        a = ws.cell(row=r, column=1).value
        if a and needle in str(a).lower():
            return r
    return None


def parse_kickoff(label):
    # Excelin kellonajat ovat Suomen aikaa; kickoff tallennetaan UTC:nä (Z).
    if not label:
        return None, None
    m = re.search(r"(\d{1,2})\.(\d{1,2})\D+klo\s+(\d{1,2}):(\d{2})", str(label), re.I)
    if not m:
        return str(label), None
    d, mo, h, mi = map(int, m.groups())
    try:
        from zoneinfo import ZoneInfo
        dt = datetime(2026, mo, d, h, mi, tzinfo=ZoneInfo("Europe/Helsinki")).astimezone(timezone.utc)
    except Exception:  # Windows ilman tzdata-pakettia: kesäturnaus = EEST (UTC+3)
        dt = datetime(2026, mo, d, h, mi, tzinfo=timezone.utc) - timedelta(hours=3)
    return str(label), dt.strftime("%Y-%m-%dT%H:%M:%SZ")


def scan_layout(ws):
    """Auto-tunnista ottelurivit, lohkot, sikajengi-/maalintekijärivit."""
    matches, groups, group_order = [], {}, []
    cur, gcount = None, 0
    for r in range(1, 92):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        gm = re.match(r"lohko\s+(\S+)\s*$", str(a).strip(), re.I) if a else None
        if gm:
            cur = gm.group(1).upper()
            groups[cur], gcount = [], 0
            group_order.append(cur)
            continue
        if b and MATCH_RE.match(str(b).strip()):
            home, away = (norm(x) for x in str(b).strip().split("-"))
            gcount += 1
            label, iso = parse_kickoff(a)
            matches.append({"id": f"{cur}{gcount}", "group": cur, "home": home,
                            "away": away, "timeLabel": label, "kickoff": iso,
                            "row": r})
            for t in (home, away):
                if t not in groups[cur]:
                    groups[cur].append(t)
    sika_row = find_label_row(ws, "Sikajengi")
    gs_row = find_label_row(ws, "Maalintekij")

    # Pisteet luetaan labelista ("Sikajengi (8p)", "Maalintekijä 2p per maali ...")
    # — säännöt voivat muuttua turnausten välillä ilman skriptimuutosta.
    def label_pts(row, pattern, default):
        if not row:
            return default
        m = re.search(pattern, str(ws.cell(row=row, column=1).value or ""))
        return int(m.group(1)) if m else default

    scoring = {
        "sikajengi": label_pts(sika_row, r"\((\d+)\s*p\)", 8),
        "goalscorer": label_pts(gs_row, r"(\d+)\s*p\s+per\s+maali", 1),
    }
    return matches, groups, group_order, {
        "sikajengi": sika_row,
        "goalscorer": gs_row,
        "scoring": scoring,
    }


def participant_cols(ws):
    out, c = [], FIRST_PRED_COL
    while True:
        name = ws.cell(row=2, column=c).value
        if name is None:
            break
        out.append((str(name).strip(), c))
        c += 2
    return out


def read_cup(ws, ranges, col):
    # Ohita placeholder-merkinnät (N/A, -) ja poista duplikaatit (ensimmäinen
    # jää) — sama joukkue kahdesti ei saa tuplata pisteitä.
    out = {}
    for key in ROUND_ORDER:
        r0, r1 = ranges[key]
        picks = [cellval(ws, r, col) for r in range(r0, r1 + 1)]
        picks = [norm(x) for x in picks if x and str(x).strip().upper() not in ("N/A", "NA", "-", "–")]
        seen, uniq = set(), []
        for x in picks:
            if x not in seen:
                seen.add(x)
                uniq.append(x)
        out[key] = (uniq[0] if uniq else None) if key == "champion" else uniq
    return out


def build_tournament(tid, matches, groups, group_order, rows):
    teams = sorted({t for ts in groups.values() for t in ts})
    sika = rows["scoring"]["sikajengi"]
    return {
        "id": tid, "name": tid.upper(), "type": "worldcup",
        "year": 2026, "startDate": "2026-06-11",
        "teams": teams,
        "groups": {g: groups[g] for g in group_order},
        "matches": [{k: v for k, v in m.items() if k != "row"} for m in matches],
        "cupRounds": [{"key": k, "label": LAYOUT["cupMeta"][k][0],
                       "pointsPerTeam": LAYOUT["cupMeta"][k][1],
                       "slots": LAYOUT["cup"][k][1] - LAYOUT["cup"][k][0] + 1}
                      for k in ROUND_ORDER],
        "scoring": {"group": {"exact": 3, "outcome": 1, "miss": 0},
                    "sikajengi": {"points": sika, "tiePoints": sika // 2},
                    "goalscorer": {"perGoal": rows["scoring"]["goalscorer"]}},
    }


def merge_tournament(new, outdir):
    """Säilytä fetch-fifa:n rikastamat kentät, jos tournament.json on jo olemassa
    (UTC-kickoffit, fifaId:t, stadionit, URL:t, teamNames, knockout). Ilman tätä
    veikkausten uudelleentuonti nollaisi otteluohjelman Excel-tasolle."""
    p = os.path.join(outdir, "tournament.json")
    if not os.path.exists(p):
        return new
    with open(p, encoding="utf-8") as f:
        old = json.load(f)
    oldm = {m["id"]: m for m in old.get("matches", [])}
    for m in new["matches"]:
        om = oldm.get(m["id"])
        if not om or om.get("home") != m["home"] or om.get("away") != m["away"]:
            continue
        for k in ("kickoff", "fifaId", "matchNumber", "stadium", "city", "url"):
            if om.get(k) is not None:
                m[k] = om[k]
        if om.get("kickoff"):
            m.pop("timeLabel", None)
    new["matches"].sort(key=lambda m: m.get("kickoff") or "")
    for k in ("teamNames", "knockout", "bracketUrl"):
        if old.get(k):
            new[k] = old[k]
    return new


def build_predictions(ws, matches, rows):
    preds = {}
    for name, c in participant_cols(ws):
        p = {"matches": {}, "cup": read_cup(ws, LAYOUT["cup"], c),
             "sikajengi": None, "goalscorer": None}
        for m in matches:
            v = cellval(ws, m["row"], c)
            if v:
                p["matches"][m["id"]] = v
        if rows["sikajengi"]:
            sg = cellval(ws, rows["sikajengi"], c)
            p["sikajengi"] = norm(sg) if sg else None
        if rows["goalscorer"]:
            gs = cellval(ws, rows["goalscorer"], c)
            p["goalscorer"] = GOALSCORER_FIX.get(gs, gs) if gs else None
        preds[name] = p
    return preds


def build_results(ws, matches, rows, existing):
    res = {"matches": {}, "dirtiestTeams": [], "rounds": {},
           "goals": (existing or {}).get("goals", {})}
    for m in matches:
        v = cellval(ws, m["row"], RESULT_COL)
        if v:
            res["matches"][m["id"]] = v
    if rows["sikajengi"]:
        sg = cellval(ws, rows["sikajengi"], RESULT_COL)
        if sg:
            res["dirtiestTeams"] = [norm(sg)]
    res["rounds"] = read_cup(ws, LAYOUT["cup"], RESULT_COL)
    return res


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("tournamentId")
    ap.add_argument("--mode", choices=["predictions", "results", "all"], default="all")
    ap.add_argument("--outdir", default=None)
    args = ap.parse_args()

    outdir = args.outdir or os.path.join("data", args.tournamentId)
    os.makedirs(outdir, exist_ok=True)
    wb = openpyxl.load_workbook(args.xlsx, data_only=False)
    if SHEET not in wb.sheetnames:
        sys.exit(f"Välilehteä '{SHEET}' ei löydy: {wb.sheetnames}")
    ws = wb[SHEET]

    matches, groups, group_order, rows = scan_layout(ws)
    if not matches:
        sys.exit("Ei löytynyt yhtään ottelua – tarkista Excel-pohja.")

    def write(name, obj):
        path = os.path.join(outdir, name)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(obj, f, ensure_ascii=False, indent=2)
        print(f"  {path}  ({len(json.dumps(obj))} tavua)")

    print(f"Excel: {args.xlsx}  →  {outdir}/")
    print(f"  {len(matches)} ottelua, {len(group_order)} lohkoa, "
          f"{len(participant_cols(ws))} osallistujaa")

    if args.mode in ("predictions", "all"):
        t = merge_tournament(build_tournament(args.tournamentId, matches, groups, group_order, rows), outdir)
        write("tournament.json", t)
        write("predictions.json", build_predictions(ws, matches, rows))
    if args.mode in ("results", "all"):
        existing = None
        rpath = os.path.join(outdir, "results.json")
        if os.path.exists(rpath):
            with open(rpath, encoding="utf-8") as f:
                existing = json.load(f)
        write("results.json", build_results(ws, matches, rows, existing))


if __name__ == "__main__":
    main()
